import csv
import openpyxl
from openpyxl import Workbook


def ordena(nomes):
    mesma_nota=0
    for x in range(0, len(nomes), 1):
        anterior = nomes[x]
        j = x - 1
        nomes[-1] = anterior
        while(float(anterior[-1]) > float(nomes[j][-1])):
            nomes[j+1] = nomes[j]
            j -= 1
        nomes[j+1] = anterior
    return nomes

def to_csv(dados, arquivo):
    wb = Workbook()
    dest_filename = arquivo

    ws = wb.active
    ws1 = wb.create_sheet(dest_filename, 0)
    ws1.title ='nome'
    print(len(dados))
    contador = 2
    ws['A1'] = 'Nome'
    ws['B1'] = 'Nota Geral'
    ws['C1'] = 'LP'
    ws['D1'] = 'Mat'
    ws['E1'] = 'CG'
    ws['F1'] = 'Inf'
    ws['G1'] = 'CE'

    for dado in dados:
        print(dado)
        nome_escrita = ""
        for nome in dado[0]:
              nome_escrita += nome + " "
        #print('adding one', a)
        #ws2.append([dado[-1]])
        celula_nome = str('A'+str(contador))
        celula_nota = str('B'+str(contador))
        celula_lp = str('C'+str(contador))
        celula_mat = str('D'+str(contador))
        celula_cg = str('E'+str(contador))
        celula_inf = str('F'+str(contador))
        celula_ce = str('G'+str(contador))
        contador += 1
        print(celula_nome)
        ws[celula_nome] = nome_escrita
        ws[celula_nota] = dado[1]
        ws[celula_lp] = dado[2]
        ws[celula_mat] = dado[3]
        ws[celula_cg] = dado[4]
        ws[celula_inf] = dado[5]
        ws[celula_ce] = dado[6]
    wb.save(filename=dest_filename)

def resultadoAux():
    notas =[]
    nomes = []
    dict_nome={}
    arquivo_ordenacao = {}
    arquivo = open("prefeitura3.txt", encoding='utf-8')
    for linha in arquivo.readlines():
        linha = linha.split(' ')
        try:
            notas.append(linha[-9])
            nNome = [linha[1:len(linha)-9], float(linha[-9]), float(linha[-5]), float(linha[-4]), float(linha[-3]),
                     float(linha[-2]), float(linha[-1])]
            nomes.append(nNome)
            dict_nome[str(linha[1:len(linha)-9])] = linha[-9]
            json ={'nome': str(linha[1:len(linha)-9]),
                   'nota': float(linha[-9])}
            arquivo_ordenacao+=json
        except:
            pass
    pessoas =ordena(nomes)
    print(pessoas)
    to_csv(pessoas, 'Auxiliar.xlsx')



resultadoAux()
