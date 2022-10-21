from openpyxl import Workbook

def ordena(nomes):
    for x in range(0, len(nomes), 1):
        anterior = nomes[x]
        j = x - 1
        nomes[-1] = anterior
        while(float(anterior[-1]) > float(nomes[j][-1])):
            nomes[j+1] = nomes[j]
            j -= 1
        nomes[j+1] = anterior
    return nomes

def to_csv(dados):
    wb = Workbook()
    dest_filename = 'Classificacaop.xlsx'
    ws = wb.active
    ws1 = wb.create_sheet(dest_filename, 0)
    ws1.title ='nome'
    print(len(dados))
    contador = 2

    ws['A1'] = 'Nome'
    ws['B1'] = 'Nota Geral'
    for dado in dados:
        nome_escrita = ""
        for nome in dado[0]:
              nome_escrita += nome + " "
        celula_nome = str('A'+str(contador))
        celula_nota = str('B'+str(contador))
        contador += 1
        print(nome_escrita)
        ws[celula_nome] = nome_escrita
        ws[celula_nota] = dado[1]
    wb.save(filename=dest_filename)

def resultadoAux():
    notas =[]
    nomes = []
    dict_nome={}
    arquivo = open("prefeiturar.txt", encoding='utf-8')
    for linha in arquivo.readlines():
        linha = linha.split(' ')
        try:
            if(float(linha[-9]) >= 00.00):
                notas.append(linha[-9])
                nNome = [linha[1:len(linha)-9], float(linha[-9])]
                nomes.append(nNome)
                dict_nome[str(linha[1:len(linha)-9])] = linha[-9]
        except:
            pass
    to_csv(ordena(nomes))

ordena(resultadoAux())